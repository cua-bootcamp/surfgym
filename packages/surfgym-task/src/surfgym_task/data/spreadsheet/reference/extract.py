#!/usr/bin/env python3

import argparse
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


def to_json_value(value: Any, *, epoch: datetime) -> Any:
    """Excel 셀 값을 SurfGym의 raw cell value와 호환되는 JSON 값으로 변환한다."""
    if value is None:
        return None

    if isinstance(value, (datetime, date, time)):
        return float(to_excel(value, epoch=epoch))

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def make_entry(
    sheet_name: str,
    coordinate: str,
    property_name: str,
    value: Any,
    *,
    epoch: datetime,
) -> dict[str, Any]:
    return {
        "spec": {
            "kind": "cell",
            "sheet": sheet_name,
            "cell": coordinate,
            "property": property_name,
        },
        "value": to_json_value(value, epoch=epoch),
    }


def to_css_rgb(color: Any) -> str | None:
    """OpenPyXL의 명시적 RGB 색상을 SurfGym이 사용하는 CSS rgb 값으로 변환한다."""
    if color is None or color.type != "rgb":
        return None

    argb = color.rgb
    if not isinstance(argb, str):
        return None

    hex_value = argb.strip()

    if len(hex_value) == 8:
        alpha = hex_value[:2]
        if alpha == "00":
            return None
        hex_value = hex_value[2:]
    elif len(hex_value) != 6:
        return None

    try:
        red = int(hex_value[0:2], 16)
        green = int(hex_value[2:4], 16)
        blue = int(hex_value[4:6], 16)
    except ValueError:
        return None

    return f"rgb({red}, {green}, {blue})"


def make_format_entries(
    sheet_name: str,
    coordinate: str,
    cell: Any,
    *,
    epoch: datetime,
) -> list[dict[str, Any]]:
    """SurfGym이 평가할 수 있는 명시적 셀 서식을 항목별로 변환한다."""
    entries: list[dict[str, Any]] = []

    if cell.number_format and cell.number_format != "General":
        entries.append(
            make_entry(
                sheet_name=sheet_name,
                coordinate=coordinate,
                property_name="numberFormat",
                value=cell.number_format,
                epoch=epoch,
            )
        )

    if cell.font.bold is True:
        entries.append(
            make_entry(
                sheet_name=sheet_name,
                coordinate=coordinate,
                property_name="bold",
                value=1,
                epoch=epoch,
            )
        )

    font_color = to_css_rgb(cell.font.color)
    if font_color is not None:
        entries.append(
            make_entry(
                sheet_name=sheet_name,
                coordinate=coordinate,
                property_name="fontColor",
                value=font_color,
                epoch=epoch,
            )
        )

    background_color = None
    if cell.fill.fill_type == "solid":
        background_color = to_css_rgb(cell.fill.fgColor)

    if background_color is not None:
        entries.append(
            make_entry(
                sheet_name=sheet_name,
                coordinate=coordinate,
                property_name="backgroundColor",
                value=background_color,
                epoch=epoch,
            )
        )

    return entries


def extract_cells(
    workbook_path: Path,
    *,
    include_formulas: bool = True,
    include_formats: bool = True,
) -> list[dict[str, Any]]:
    """XLSX의 셀 값과 SurfGym이 지원하는 수식/서식 정보를 추출한다."""
    # 수식 자체를 읽는 workbook
    formula_workbook = load_workbook(
        workbook_path,
        data_only=False,
        read_only=False,
    )

    # 수식의 저장된 계산 결과를 읽는 workbook
    value_workbook = load_workbook(
        workbook_path,
        data_only=True,
        read_only=False,
    )

    entries: list[dict[str, Any]] = []

    try:
        epoch = value_workbook.epoch

        for formula_sheet in formula_workbook.worksheets:
            sheet_name = formula_sheet.title
            value_sheet = value_workbook[sheet_name]

            for row in formula_sheet.iter_rows():
                for formula_cell in row:
                    coordinate = formula_cell.coordinate
                    raw_value = formula_cell.value

                    # 기존 추출 범위와 동일하게 값이 없는 셀은 제외한다.
                    if raw_value is None:
                        continue

                    format_entries = (
                        make_format_entries(
                            sheet_name=sheet_name,
                            coordinate=coordinate,
                            cell=formula_cell,
                            epoch=epoch,
                        )
                        if include_formats
                        else []
                    )

                    if formula_cell.data_type == "f":
                        if include_formulas:
                            entries.append(
                                make_entry(
                                    sheet_name=sheet_name,
                                    coordinate=coordinate,
                                    property_name="formula",
                                    value=raw_value,
                                    epoch=epoch,
                                )
                            )

                        # XLSX에 마지막으로 저장된 수식 계산 결과
                        entries.append(
                            make_entry(
                                sheet_name=sheet_name,
                                coordinate=coordinate,
                                property_name="value",
                                value=value_sheet[coordinate].value,
                                epoch=epoch,
                            )
                        )
                    else:
                        entries.append(
                            make_entry(
                                sheet_name=sheet_name,
                                coordinate=coordinate,
                                property_name="value",
                                value=raw_value,
                                epoch=epoch,
                            )
                        )

                    entries.extend(format_entries)
    finally:
        formula_workbook.close()
        value_workbook.close()

    return entries


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "XLSX 파일의 셀 값, 수식, SurfGym 평가 가능 서식을 "
            "JSON 배열로 추출합니다."
        )
    )
    parser.add_argument(
        "input",
        type=Path,
        help="입력 XLSX/XLSM 파일",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="출력 JSON 파일. 생략하면 입력 파일명_cells.json을 사용합니다.",
    )
    formula_group = parser.add_mutually_exclusive_group()
    formula_group.add_argument(
        "--include-formulas",
        dest="include_formulas",
        action="store_true",
        help="수식 셀의 formula 항목을 추출합니다. 기본값입니다.",
    )
    formula_group.add_argument(
        "--no-formulas",
        dest="include_formulas",
        action="store_false",
        help="수식 셀의 formula 항목을 제외합니다.",
    )

    format_group = parser.add_mutually_exclusive_group()
    format_group.add_argument(
        "--include-formats",
        dest="include_formats",
        action="store_true",
        help="지원되는 셀 서식 항목을 추출합니다. 기본값입니다.",
    )
    format_group.add_argument(
        "--no-formats",
        dest="include_formats",
        action="store_false",
        help="셀 서식 항목을 제외합니다.",
    )
    parser.set_defaults(
        include_formulas=True,
        include_formats=True,
    )
    args = parser.parse_args()

    input_path: Path = args.input

    if not input_path.exists():
        raise FileNotFoundError(f"파일이 없습니다: {input_path}")

    output_path = args.output or input_path.with_name(f"{input_path.stem}_cells.json")

    entries = extract_cells(
        input_path,
        include_formulas=args.include_formulas,
        include_formats=args.include_formats,
    )

    output_path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"추출한 항목: {len(entries)}개")
    print(f"출력 파일: {output_path.resolve()}")


if __name__ == "__main__":
    main()
